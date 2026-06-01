"""
用户信息备份插件 (Pro 专享)
支持导出用户信息到 Excel，支持本地备份、WebDAV、下载
支持增量恢复和覆盖恢复，支持选择性恢复
支持定时自动备份
"""
import os
import json
import logging
import hashlib
import time
import threading
from datetime import datetime, timedelta
from io import BytesIO
from typing import List, Dict, Optional, Any
from fastapi import Request, UploadFile, File
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.plugins.user_backup.user_backup_dao import (
    get_user_meta_for_backup,
    list_point_logs_for_backup,
    list_tg_bindings_detail_for_backup,
    list_users_meta_for_backup,
    replace_point_logs_for_backup,
    upsert_user_meta_for_backup,
)
from app.plugins.base import PluginBase
from app.domains.users.auth import is_admin_user  # 🔒 管理员鉴权
from app.infra.clients.media_server_client import media_api
from app.infra.clients.webdav_client import webdav_client
from app.infra.config.media_server_settings import get_media_server_host

logger = logging.getLogger("uvicorn")

# 备份存储目录
BACKUP_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "data", "user_backups")

# 可恢复的字段列表
RESTORE_FIELDS = [
    {"key": "expire_date", "label": "过期时间", "source": "meta"},
    {"key": "is_vip", "label": "VIP状态", "source": "meta"},
    {"key": "points", "label": "积分余额", "source": "meta"},
    {"key": "max_concurrent", "label": "并发限制", "source": "meta"},
    {"key": "allow_routes", "label": "允许线路", "source": "meta"},
    {"key": "block_routes", "label": "屏蔽线路", "source": "meta"},
    {"key": "remark", "label": "备注", "source": "meta"},
    {"key": "risk_level", "label": "风险等级", "source": "meta"},
    {"key": "req_free", "label": "求片权限", "source": "meta"},
    {"key": "req_free_count", "label": "免费求片次数", "source": "meta"},
    {"key": "is_disabled", "label": "禁用状态", "source": "emby"},
    {"key": "enable_all_folders", "label": "全部媒体库", "source": "emby"},
    {"key": "enabled_folders", "label": "可用媒体库", "source": "emby"},
    {"key": "enable_downloading", "label": "下载权限", "source": "emby"},
    {"key": "enable_video_transcoding", "label": "视频转码", "source": "emby"},
    {"key": "enable_audio_transcoding", "label": "音频转码", "source": "emby"},
    {"key": "max_parental_rating", "label": "家长控制", "source": "emby"},
]


class UserBackupPlugin(PluginBase):
    id = "user_backup"
    name = "用户信息备份"
    description = "导出用户信息到Excel，支持本地/WebDAV备份，增量/覆盖恢复，定时自动备份（Pro 专享）"
    icon = "fa-database"
    icon_color = "from-blue-500 to-cyan-500"
    version = "1.1.0"
    author = "EmbyPulse"
    pro_only = True

    def __init__(self):
        super().__init__()
        self._thread = None
        self._running = False
        self._last_backup_date = None
        self._webdav_dir_created = False  # WebDAV 目录是否已创建标记
        self._setup_routes()
        self._ensure_dir()

    def _ensure_dir(self):
        """确保备份目录存在"""
        os.makedirs(BACKUP_DIR, exist_ok=True)

    def _is_pro(self) -> bool:
        """检查 Pro 授权"""
        return True

    def _get_webdav_config(self) -> Dict:
        """获取 WebDAV 配置"""
        config = self._get_config()
        return {
            "enabled": config.get("webdav_enabled", False),
            "url": config.get("webdav_url", ""),
            "user": config.get("webdav_user", ""),
            "password": config.get("webdav_password", ""),
            "path": config.get("webdav_path", "/backups/embypulse/")
        }

    def _upload_to_webdav(self, filename: str, file_data: bytes) -> bool:
        """上传文件到 WebDAV"""
        webdav = self._get_webdav_config()
        if not webdav["enabled"] or not webdav["url"]:
            return False
        
        try:
            base_url = webdav["url"].rstrip("/")
            path = webdav["path"].lstrip("/")
            
            # 确保路径以 / 结尾
            if not path.endswith("/"):
                path += "/"
            
            # 先尝试创建目录（如果不存在）
            dir_url = base_url + "/" + path
            auth = (webdav["user"], webdav["password"]) if webdav["user"] else None
            
            # 使用 MKCOL 创建目录（只在首次创建时记录日志）
            if not self._webdav_dir_created:
                try:
                    mkcol_resp = webdav_client.request('MKCOL', dir_url.rstrip("/"), auth=auth, timeout=10)
                    if mkcol_resp.status_code in [200, 201, 204]:
                        self.log(f"📁 WebDAV 创建目录成功: {path}")
                        self._webdav_dir_created = True
                    # 405 表示目录已存在，409 也可能表示已存在或路径冲突
                    elif mkcol_resp.status_code in [405, 409]:
                        self._webdav_dir_created = True  # 目录已存在，标记为已创建
                    else:
                        self.log(f"⚠️ WebDAV 目录状态: HTTP {mkcol_resp.status_code}")
                except Exception as e:
                    self.log(f"⚠️ WebDAV 创建目录异常: {e}")
            
            # 上传文件
            file_url = dir_url + filename
            resp = webdav_client.put(file_url, data=file_data, auth=auth, timeout=30)
            
            if resp.status_code in [200, 201, 204]:
                self.log(f"✅ WebDAV 上传成功: {filename}")
                return True
            elif resp.status_code == 409:
                # 409 Conflict - 可能是目录不存在，再次尝试创建后上传
                self.log(f"⚠️ WebDAV 409 冲突，尝试重新创建目录")
                webdav_client.request('MKCOL', dir_url.rstrip("/"), auth=auth, timeout=10)
                retry_resp = webdav_client.put(file_url, data=file_data, auth=auth, timeout=30)
                if retry_resp.status_code in [200, 201, 204]:
                    self.log(f"✅ WebDAV 重试上传成功: {filename}")
                    return True
                self.log(f"❌ WebDAV 重试上传失败: {retry_resp.status_code}")
                return False
            else:
                self.log(f"❌ WebDAV 上传失败: HTTP {resp.status_code}")
                return False
        except Exception as e:
            self.log(f"❌ WebDAV 上传异常: {e}")
            return False

    def _collect_user_data(self) -> List[Dict]:
        """收集所有用户数据"""
        try:
            # 获取 Emby 用户列表
            res = media_api.get("/Users", timeout=10)
            if res.status_code != 200:
                return []
            
            emby_users = res.json()
            
            # 获取本地扩展属性
            meta_rows = list_users_meta_for_backup()
            meta_map = {r['user_id']: dict(r) for r in meta_rows} if meta_rows else {}
            
            # 获取 TG 绑定关系
            tg_bindings = {}
            try:
                rows = list_tg_bindings_detail_for_backup()
                tg_bindings = {row["emby_user_id"]: row["tg_user_id"] for row in rows if row["emby_user_id"]}
            except Exception:
                pass
            
            users = []
            for u in emby_users:
                uid = u['Id']
                policy = u.get('Policy', {})
                meta = meta_map.get(uid, {})
                
                users.append({
                    "user_id": uid,
                    "name": u['Name'],
                    "is_disabled": policy.get('IsDisabled', False),
                    "is_admin": policy.get('IsAdministrator', False),
                    "expire_date": meta.get('expire_date', ''),
                    "is_vip": bool(meta.get('is_vip', 0)),
                    "points": meta.get('points', 0) or 0,
                    "max_concurrent": meta.get('max_concurrent', ''),
                    "allow_routes": meta.get('allow_routes', ''),
                    "block_routes": meta.get('block_routes', ''),
                    "remark": meta.get('remark', ''),
                    "risk_level": meta.get('risk_level', 'safe'),
                    "req_free": meta.get('req_free', 0) or 0,
                    "req_free_count": meta.get('req_free_count') if meta.get('req_free_count') is not None else -1,
                    "admin_disabled": bool(meta.get('admin_disabled', 0)),
                    "enable_all_folders": policy.get('EnableAllFolders', True),
                    "enabled_folders": ','.join(policy.get('EnabledFolders', [])),
                    "enable_downloading": policy.get('EnableContentDownloading', True),
                    "enable_video_transcoding": policy.get('EnableVideoPlaybackTranscoding', True),
                    "enable_audio_transcoding": policy.get('EnableAudioPlaybackTranscoding', True),
                    "max_parental_rating": policy.get('MaxParentalRating', ''),
                    "tg_user_id": tg_bindings.get(uid, ''),
                    "last_login": u.get('LastLoginDate', '')
                })
            
            return users
        except Exception as e:
            logger.error(f"[用户备份] 收集用户数据失败: {e}")
            return []

    def _collect_point_logs(self) -> List[Dict]:
        """收集积分变动记录"""
        try:
            rows = list_point_logs_for_backup()
            return [dict(r) for r in rows] if rows else []
        except:
            return []

    def _collect_tg_bindings(self) -> List[Dict]:
        """收集 TG 绑定关系"""
        try:
            rows = list_tg_bindings_detail_for_backup()
            return [
                {
                    "tg_user_id": r["tg_user_id"],
                    "emby_user_id": r["emby_user_id"],
                    "emby_username": r["emby_username"],
                    "bound_at": r["bound_at"] if r["bound_at"] else '',
                }
                for r in rows
            ]
        except:
            return []

    def _create_excel(self) -> bytes:
        """创建 Excel 备份文件"""
        wb = Workbook()
        
        # 样式定义
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: 用户信息
        ws1 = wb.active
        ws1.title = "用户信息"
        
        headers = [
            "用户ID", "用户名", "过期时间", "VIP", "积分", "并发限制", 
            "允许线路", "屏蔽线路", "备注", "风险等级", "求片权限", "免费求片次数",
            "禁用状态", 
            "全部媒体库", "可用媒体库", "下载权限", "视频转码", "音频转码",
            "家长控制", "TG绑定ID", "最后登录", "备份时间"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        
        users = self._collect_user_data()
        backup_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for row_idx, user in enumerate(users, 2):
            # 求片权限显示文本
            req_free_text = "免费求片" if user['req_free'] == 1 else "跟随全局"
            req_free_count_text = ("无限" if user['req_free_count'] == -1 else str(user['req_free_count'])) if user['req_free'] == 1 else ""
            
            values = [
                user['user_id'],
                user['name'],
                user['expire_date'] or '',
                "是" if user['is_vip'] else "否",
                user['points'],
                user['max_concurrent'] or '',
                user['allow_routes'] or '',
                user['block_routes'] or '',
                user['remark'] or '',
                user['risk_level'] or 'safe',
                req_free_text,
                req_free_count_text,
                "是" if user['is_disabled'] else "否",
                "是" if user['enable_all_folders'] else "否",
                user['enabled_folders'] or '',
                "是" if user['enable_downloading'] else "否",
                "是" if user['enable_video_transcoding'] else "否",
                "是" if user['enable_audio_transcoding'] else "否",
                user['max_parental_rating'] or '',
                user['tg_user_id'] or '',
                user['last_login'] or '',
                backup_time
            ]
            for col, value in enumerate(values, 1):
                cell = ws1.cell(row=row_idx, column=col, value=value)
                cell.border = border
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            ws1.column_dimensions[get_column_letter(col)].width = 15
        
        # Sheet 2: TG绑定关系
        ws2 = wb.create_sheet("TG绑定关系")
        tg_headers = ["TG用户ID", "Emby用户ID", "Emby用户名", "绑定时间"]
        for col, header in enumerate(tg_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        tg_bindings = self._collect_tg_bindings()
        for row_idx, binding in enumerate(tg_bindings, 2):
            ws2.cell(row=row_idx, column=1, value=binding['tg_user_id'])
            ws2.cell(row=row_idx, column=2, value=binding['emby_user_id'])
            ws2.cell(row=row_idx, column=3, value=binding['emby_username'])
            ws2.cell(row=row_idx, column=4, value=binding['bound_at'] or '')
        
        # Sheet 3: 积分记录
        ws3 = wb.create_sheet("积分记录")
        point_headers = ["ID", "用户ID", "用户名", "操作", "数额", "余额", "时间"]
        for col, header in enumerate(point_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        point_logs = self._collect_point_logs()
        for row_idx, log in enumerate(point_logs, 2):
            ws3.cell(row=row_idx, column=1, value=log.get('id', ''))
            ws3.cell(row=row_idx, column=2, value=log.get('user_id', ''))
            ws3.cell(row=row_idx, column=3, value=log.get('username', ''))
            ws3.cell(row=row_idx, column=4, value=log.get('action', ''))
            ws3.cell(row=row_idx, column=5, value=log.get('amount', 0))
            ws3.cell(row=row_idx, column=6, value=log.get('balance', 0))
            ws3.cell(row=row_idx, column=7, value=log.get('created_at', ''))
        
        # Sheet 4: 备份元信息
        ws4 = wb.create_sheet("备份元信息")
        meta_info = [
            ("备份版本", "1.0"),
            ("备份时间", backup_time),
            ("用户总数", len(users)),
            ("TG绑定数", len(tg_bindings)),
            ("积分记录数", len(point_logs)),
            ("Emby服务器", get_media_server_host()),
            ("Pro状态", "是" if self._is_pro() else "否"),
        ]
        for row_idx, (key, value) in enumerate(meta_info, 1):
            ws4.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            ws4.cell(row=row_idx, column=2, value=value)
        
        # 保存到内存
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _setup_routes(self):
        """注册插件 API 路由"""

        @self.router.get("/status")
        async def get_status(request: Request):
            """获取插件状态"""
            return {
                "status": "success",
                "data": {
                    "is_pro": self._is_pro(),
                    "backup_dir": BACKUP_DIR,
                    "restore_fields": RESTORE_FIELDS
                }
            }

        @self.router.post("/create")
        async def create_backup(request: Request):
            """创建备份"""
            if not request.session.get("user"):
                return {"status": "error", "message": "未登录"}
            if not is_admin_user(request):
                return {"status": "error", "message": "需要管理员权限"}
            
            if not self._is_pro():
                return {"status": "error", "message": "此功能需要 Pro 授权", "need_pro": True}
            
            try:
                # 创建 Excel
                file_data = self._create_excel()
                
                # 生成文件名
                filename = f"user_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                filepath = os.path.join(BACKUP_DIR, filename)
                
                # 保存本地
                with open(filepath, "wb") as f:
                    f.write(file_data)
                
                # 上传 WebDAV
                webdav_ok = self._upload_to_webdav(filename, file_data)
                
                # 清理旧备份
                self._cleanup_old_backups()
                
                file_size = len(file_data)
                md5 = hashlib.md5(file_data).hexdigest()
                
                self.log(f"✅ 创建备份: {filename} ({file_size} bytes)")
                
                return {
                    "status": "success",
                    "data": {
                        "filename": filename,
                        "size": file_size,
                        "md5": md5,
                        "webdav_uploaded": webdav_ok,
                        "download_url": f"/api/plugins/user_backup/download/{filename}"
                    },
                    "message": "备份创建成功"
                }
            except Exception as e:
                logger.error(f"[用户备份] 创建失败: {e}")
                return {"status": "error", "message": str(e)}

        @self.router.get("/list")
        async def list_backups(request: Request):
            """获取备份列表"""
            if not request.session.get("user"):
                return {"status": "error", "message": "未登录"}
            if not is_admin_user(request):
                return {"status": "error", "message": "需要管理员权限"}
            
            try:
                backups = []
                if os.path.exists(BACKUP_DIR):
                    for f in sorted(os.listdir(BACKUP_DIR), reverse=True):
                        if f.endswith(".xlsx") and f.startswith("user_backup_"):
                            filepath = os.path.join(BACKUP_DIR, f)
                            stat = os.stat(filepath)
                            
                            # 解析备份时间
                            try:
                                time_str = f.replace("user_backup_", "").replace(".xlsx", "")
                                backup_time = datetime.strptime(time_str, "%Y%m%d_%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
                            except:
                                backup_time = ""
                            
                            backups.append({
                                "filename": f,
                                "size": stat.st_size,
                                "created_at": backup_time,
                                "download_url": f"/api/plugins/user_backup/download/{f}"
                            })
                
                return {"status": "success", "data": backups}
            except Exception as e:
                return {"status": "error", "message": str(e)}

        @self.router.get("/download/{filename}")
        async def download_backup(filename: str, request: Request):
            """下载备份文件"""
            if not request.session.get("user"):
                return {"status": "error", "message": "未登录"}
            if not is_admin_user(request):
                return {"status": "error", "message": "需要管理员权限"}
            
            # 安全检查
            if ".." in filename or "/" in filename or "\\" in filename:
                return {"status": "error", "message": "无效文件名"}
            
            filepath = os.path.join(BACKUP_DIR, filename)
            if not os.path.exists(filepath):
                return {"status": "error", "message": "文件不存在"}
            
            from fastapi.responses import FileResponse
            return FileResponse(
                filepath,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=filename
            )

        @self.router.delete("/delete/{filename}")
        async def delete_backup(filename: str, request: Request):
            """删除备份"""
            if not request.session.get("user"):
                return {"status": "error", "message": "未登录"}
            if not is_admin_user(request):
                return {"status": "error", "message": "需要管理员权限"}
            
            if not self._is_pro():
                return {"status": "error", "message": "此功能需要 Pro 授权", "need_pro": True}
            
            # 安全检查
            if ".." in filename or "/" in filename or "\\" in filename:
                return {"status": "error", "message": "无效文件名"}
            
            filepath = os.path.join(BACKUP_DIR, filename)
            if os.path.exists(filepath):
                os.remove(filepath)
                self.log(f"🗑️ 删除备份: {filename}")
            
            return {"status": "success", "message": "删除成功"}

        @self.router.post("/upload")
        async def upload_backup_for_restore(request: Request, file: UploadFile = File(...)):
            """上传备份文件用于恢复"""
            if not request.session.get("user"):
                return {"status": "error", "message": "未登录"}
            if not is_admin_user(request):
                return {"status": "error", "message": "需要管理员权限"}
            
            if not self._is_pro():
                return {"status": "error", "message": "此功能需要 Pro 授权", "need_pro": True}
            
            try:
                content = await file.read()
                if not content:
                    return {"status": "error", "message": "文件内容为空"}
                
                # 解析 Excel
                wb = load_workbook(BytesIO(content))
                
                # 读取用户信息 Sheet
                ws = wb["用户信息"]
                users = []
                headers = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    if row[0].value:  # 有用户ID
                        user_data = {}
                        for col_idx, cell in enumerate(row):
                            key = headers[col_idx] if col_idx < len(headers) else ""
                            if key:
                                user_data[key] = cell.value
                        users.append(user_data)
                
                # 读取积分记录 Sheet
                point_logs = []
                if "积分记录" in wb.sheetnames:
                    ws_points = wb["积分记录"]
                    point_headers = [cell.value for cell in ws_points[1]]
                    for row in ws_points.iter_rows(min_row=2, max_row=ws_points.max_row):
                        if row[0].value:
                            log_data = {}
                            for col_idx, cell in enumerate(row):
                                key = point_headers[col_idx] if col_idx < len(point_headers) else ""
                                if key:
                                    log_data[key] = cell.value
                            point_logs.append(log_data)
                
                # 保存临时文件
                temp_filename = f"temp_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                temp_filepath = os.path.join(BACKUP_DIR, temp_filename)
                with open(temp_filepath, "wb") as f:
                    f.write(content)
                
                return {
                    "status": "success",
                    "data": {
                        "temp_filename": temp_filename,
                        "user_count": len(users),
                        "point_log_count": len(point_logs),
                        "users": users[:20],  # 返回前20条预览
                        "restore_fields": RESTORE_FIELDS
                    },
                    "message": f"文件解析成功，共 {len(users)} 个用户"
                }
            except Exception as e:
                logger.error(f"[用户备份] 上传解析失败: {e}")
                return {"status": "error", "message": f"文件解析失败: {str(e)}"}

        @self.router.post("/preview")
        async def preview_restore(request: Request):
            """预览恢复差异"""
            if not request.session.get("user"):
                return {"status": "error", "message": "未登录"}
            if not is_admin_user(request):
                return {"status": "error", "message": "需要管理员权限"}
            
            if not self._is_pro():
                return {"status": "error", "message": "此功能需要 Pro 授权", "need_pro": True}
            
            try:
                data = await request.json()
                filename = data.get("filename")
                fields = data.get("fields", [])  # 要预览的字段列表
                
                if not filename:
                    return {"status": "error", "message": "缺少文件名"}
                
                filepath = os.path.join(BACKUP_DIR, filename)
                if not os.path.exists(filepath):
                    return {"status": "error", "message": "文件不存在"}
                
                # 解析备份文件
                wb = load_workbook(filepath)
                ws = wb["用户信息"]
                headers = [cell.value for cell in ws[1]]
                
                backup_users = {}
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    if row[0].value:
                        user_data = {}
                        for col_idx, cell in enumerate(row):
                            key = headers[col_idx] if col_idx < len(headers) else ""
                            if key:
                                user_data[key] = cell.value
                        backup_users[user_data["用户ID"]] = user_data
                
                # 获取当前用户数据
                current_users = self._collect_user_data()
                current_map = {u['user_id']: u for u in current_users}
                
                # 比较差异
                diff_list = []
                for uid, backup_user in backup_users.items():
                    current_user = current_map.get(uid, {})
                    
                    # 标记已删除的用户
                    is_deleted = not bool(current_user)
                    
                    diff = {
                        "user_id": uid,
                        "name": backup_user.get("用户名", ""),
                        "exists_local": bool(current_user),
                        "is_deleted": is_deleted,  # 新增：标记已删除
                        "changes": []
                    }
                    
                    # 已删除的用户不需要显示差异，直接标记
                    if is_deleted:
                        diff["changes"] = [{"field": "status", "field_label": "状态", "backup_value": "备份中存在", "current_value": "已删除"}]
                        diff_list.append(diff)
                        continue
                    
                    # 字段映射
                    field_map = {
                        "expire_date": ("过期时间", "expire_date"),
                        "is_vip": ("VIP", "is_vip"),
                        "points": ("积分", "points"),
                        "max_concurrent": ("并发限制", "max_concurrent"),
                        "allow_routes": ("允许线路", "allow_routes"),
                        "block_routes": ("屏蔽线路", "block_routes"),
                        "remark": ("备注", "remark"),
                        "risk_level": ("风险等级", "risk_level"),
                        "req_free": ("求片权限", "req_free"),
                        "req_free_count": ("免费求片次数", "req_free_count"),
                        "is_disabled": ("禁用状态", "is_disabled"),
                        "enable_all_folders": ("全部媒体库", "enable_all_folders"),
                        "enabled_folders": ("可用媒体库", "enabled_folders"),
                        "enable_downloading": ("下载权限", "enable_downloading"),
                        "enable_video_transcoding": ("视频转码", "enable_video_transcoding"),
                        "enable_audio_transcoding": ("音频转码", "enable_audio_transcoding"),
                        "max_parental_rating": ("家长控制", "max_parental_rating"),
                    }
                    
                    for field_key in fields:
                        if field_key in field_map:
                            backup_key, local_key = field_map[field_key]
                            backup_val = backup_user.get(backup_key, "")
                            current_val = current_user.get(local_key, "") if current_user else ""
                            
                            # 标准化值进行比较
                            # 布尔字段：转为 "是"/"否"
                            if backup_key in ["VIP", "禁用状态", "全部媒体库", "下载权限", "视频转码", "音频转码"]:
                                # 备份值：Excel 中是 "是"/"否" 字符串，或者 True/False
                                if backup_val is True or backup_val == "是":
                                    backup_val = "是"
                                else:
                                    backup_val = "否"
                                # 当前值：Python bool
                                if current_val is True:
                                    current_val = "是"
                                else:
                                    current_val = "否"
                            
                            # 求片权限字段：转为文本比较
                            elif backup_key == "求片权限":
                                # 备份值：Excel 中是 "免费求片"/"跟随全局"
                                backup_val = str(backup_val or "跟随全局")
                                # 当前值：0=跟随全局, 1=免费求片
                                if current_val == 1:
                                    current_val = "免费求片"
                                else:
                                    current_val = "跟随全局"
                            
                            # 免费求片次数字段：转为文本比较
                            elif backup_key == "免费求片次数":
                                # 备份值：Excel 中是 "无限" 或数字文本
                                if backup_val == "无限" or backup_val == -1:
                                    backup_val = "无限"
                                else:
                                    backup_val = str(backup_val) if backup_val not in (None, "") else ""
                                # 当前值：-1=无限, >=0=次数
                                if current_val == -1 or current_val is None:
                                    current_val = "无限"
                                else:
                                    current_val = str(current_val)
                            
                            # 数字字段：转为字符串比较
                            elif backup_key in ["积分", "家长控制"]:
                                backup_val = str(backup_val or "")
                                current_val = str(current_val or "")
                            
                            # 其他字段：空值统一为空字符串
                            else:
                                backup_val = str(backup_val or "").strip()
                                current_val = str(current_val or "").strip()
                            
                            # 只有值真正不同才记录差异
                            if backup_val != current_val:
                                diff["changes"].append({
                                    "field": field_key,
                                    "field_label": backup_key,
                                    "backup_value": backup_val if backup_val else "(空)",
                                    "current_value": current_val if current_val else "(空)"
                                })
                    
                    # 只记录有差异的用户
                    if diff["changes"]:
                        diff_list.append(diff)
                
                return {
                    "status": "success",
                    "data": {
                        "diff_count": len(diff_list),
                        "diffs": diff_list[:50]  # 返回前50条
                    }
                }
            except Exception as e:
                logger.error(f"[用户备份] 预览失败: {e}")
                return {"status": "error", "message": str(e)}

        @self.router.post("/restore")
        async def restore_backup(request: Request):
            """执行恢复"""
            if not request.session.get("user"):
                return {"status": "error", "message": "未登录"}
            if not is_admin_user(request):
                return {"status": "error", "message": "需要管理员权限"}
            
            if not self._is_pro():
                return {"status": "error", "message": "此功能需要 Pro 授权", "need_pro": True}
            
            try:
                data = await request.json()
                filename = data.get("filename")
                mode = data.get("mode", "increment")  # increment / overwrite
                user_ids = data.get("user_ids", [])  # 选择性恢复的用户列表，空=全部
                fields = data.get("fields", [])  # 选择性恢复的字段列表
                
                if not filename:
                    return {"status": "error", "message": "缺少文件名"}
                
                filepath = os.path.join(BACKUP_DIR, filename)
                if not os.path.exists(filepath):
                    return {"status": "error", "message": "文件不存在"}
                
                # 解析备份文件
                wb = load_workbook(filepath)
                ws = wb["用户信息"]
                headers = [cell.value for cell in ws[1]]
                
                backup_users = []
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    if row[0].value:
                        user_data = {}
                        for col_idx, cell in enumerate(row):
                            key = headers[col_idx] if col_idx < len(headers) else ""
                            if key:
                                user_data[key] = cell.value
                        backup_users.append(user_data)
                
                # 积分记录
                point_logs = []
                if "积分记录" in wb.sheetnames and "points" in fields:
                    ws_points = wb["积分记录"]
                    for row in ws_points.iter_rows(min_row=2, max_row=ws_points.max_row):
                        if row[0].value:
                            point_logs.append({
                                "id": row[0].value,
                                "user_id": row[1].value,
                                "username": row[2].value,
                                "action": row[3].value,
                                "amount": row[4].value,
                                "balance": row[5].value,
                                "created_at": row[6].value
                            })
                
                # 执行恢复
                success_count = 0
                fail_count = 0
                results = []
                
                for backup_user in backup_users:
                    uid = backup_user.get("用户ID")
                    uname = backup_user.get("用户名", "")
                    
                    # 选择性恢复：跳过未选中的用户
                    if user_ids and uid not in user_ids:
                        continue
                    
                    try:
                        # 获取当前用户状态
                        res = media_api.get(f"/Users/{uid}", timeout=5)
                        user_exists = res.status_code == 200
                        
                        # 增量模式：跳过本地不存在的用户
                        if mode == "increment" and not user_exists:
                            continue
                        
                        current_policy = {}
                        if user_exists:
                            current_policy = res.json().get('Policy', {})
                        
                        # 获取当前 meta
                        current_meta = get_user_meta_for_backup(uid)
                        
                        # 更新 meta 字段
                        meta_updates = {}
                        
                        if "expire_date" in fields:
                            backup_val = backup_user.get("过期时间", "")
                            current_val = current_meta.get('expire_date') if current_meta else None
                            # 增量模式：值相同则跳过（备份空值也能覆盖本地值）
                            if mode == "increment" and str(current_val or "") == str(backup_val or ""):
                                pass
                            else:
                                meta_updates['expire_date'] = backup_val if backup_val else None
                        
                        if "is_vip" in fields:
                            backup_val = backup_user.get("VIP", "否") == "是"
                            current_val = bool(current_meta.get('is_vip')) if current_meta else False
                            # 增量模式：值相同则跳过
                            if mode == "increment" and current_val == backup_val:
                                pass
                            else:
                                meta_updates['is_vip'] = 1 if backup_val else 0
                        
                        if "points" in fields:
                            backup_val = int(backup_user.get("积分", 0) or 0)
                            current_val = int(current_meta.get('points') or 0) if current_meta else 0
                            # 增量模式：值相同则跳过
                            if mode == "increment" and current_val == backup_val:
                                pass
                            else:
                                meta_updates['points'] = backup_val
                        
                        if "max_concurrent" in fields:
                            backup_val = backup_user.get("并发限制", "")
                            current_val = current_meta.get('max_concurrent') if current_meta else None
                            # 增量模式：值相同则跳过
                            if mode == "increment" and str(current_val or "") == str(backup_val or ""):
                                pass
                            else:
                                meta_updates['max_concurrent'] = backup_val if backup_val else None
                        
                        if "allow_routes" in fields:
                            backup_val = backup_user.get("允许线路", "")
                            current_val = current_meta.get('allow_routes') if current_meta else None
                            # 增量模式：值相同则跳过
                            if mode == "increment" and str(current_val or "") == str(backup_val or ""):
                                pass
                            else:
                                meta_updates['allow_routes'] = backup_val
                        
                        if "block_routes" in fields:
                            backup_val = backup_user.get("屏蔽线路", "")
                            current_val = current_meta.get('block_routes') if current_meta else None
                            # 增量模式：值相同则跳过
                            if mode == "increment" and str(current_val or "") == str(backup_val or ""):
                                pass
                            else:
                                meta_updates['block_routes'] = backup_val
                        
                        if "remark" in fields:
                            backup_val = backup_user.get("备注", "")
                            current_val = current_meta.get('remark') if current_meta else None
                            # 增量模式：值相同则跳过
                            if mode == "increment" and str(current_val or "") == str(backup_val or ""):
                                pass
                            else:
                                meta_updates['remark'] = backup_val
                        
                        if "risk_level" in fields:
                            backup_val = backup_user.get("风险等级", "safe")
                            current_val = current_meta.get('risk_level') if current_meta else "safe"
                            # 增量模式：值相同则跳过
                            if mode == "increment" and str(current_val or "safe") == str(backup_val or "safe"):
                                pass
                            else:
                                meta_updates['risk_level'] = backup_val
                        
                        if "req_free" in fields:
                            backup_val_text = backup_user.get("求片权限", "跟随全局")
                            backup_val = 1 if backup_val_text == "免费求片" else 0
                            current_val = int(current_meta.get('req_free') or 0) if current_meta else 0
                            # 增量模式：值相同则跳过
                            if mode == "increment" and current_val == backup_val:
                                pass
                            else:
                                meta_updates['req_free'] = backup_val
                        
                        if "req_free_count" in fields:
                            backup_val_text = backup_user.get("免费求片次数", "无限")
                            if backup_val_text == "无限":
                                backup_val = -1
                            else:
                                backup_val = int(backup_val_text) if backup_val_text not in (None, "") else -1
                            current_val = int(current_meta.get('req_free_count')) if current_meta and current_meta.get('req_free_count') is not None else -1
                            # 增量模式：值相同则跳过
                            if mode == "increment" and current_val == backup_val:
                                pass
                            else:
                                meta_updates['req_free_count'] = backup_val
                        
                        # 更新 Policy 字段
                        policy_updates = {}
                        
                        if "is_disabled" in fields:
                            val = backup_user.get("禁用状态", "否") == "是"
                            if mode == "increment":
                                # 增量模式只恢复启用状态，不恢复禁用
                                if not val and current_policy.get('IsDisabled'):
                                    policy_updates['IsDisabled'] = False
                            else:
                                policy_updates['IsDisabled'] = val
                        
                        if "enable_all_folders" in fields:
                            val = backup_user.get("全部媒体库", "是") == "是"
                            policy_updates['EnableAllFolders'] = val
                        
                        if "enabled_folders" in fields:
                            val = backup_user.get("可用媒体库", "")
                            if val:
                                folders = [f.strip() for f in val.split(',') if f.strip()]
                                policy_updates['EnabledFolders'] = folders
                        
                        if "enable_downloading" in fields:
                            val = backup_user.get("下载权限", "是") == "是"
                            policy_updates['EnableContentDownloading'] = val
                            policy_updates['EnableSyncTranscoding'] = val
                        
                        if "enable_video_transcoding" in fields:
                            val = backup_user.get("视频转码", "是") == "是"
                            policy_updates['EnableVideoPlaybackTranscoding'] = val
                            policy_updates['EnablePlaybackRemuxing'] = val
                        
                        if "enable_audio_transcoding" in fields:
                            val = backup_user.get("音频转码", "是") == "是"
                            policy_updates['EnableAudioPlaybackTranscoding'] = val
                        
                        if "max_parental_rating" in fields:
                            val = backup_user.get("家长控制", "")
                            if val:
                                policy_updates['MaxParentalRating'] = int(val)
                            elif mode == "overwrite":
                                policy_updates.pop('MaxParentalRating', None)
                        
                        # 执行更新
                        if meta_updates:
                            if current_meta:
                                upsert_user_meta_for_backup(uid, meta_updates, datetime.now().isoformat())
                            else:
                                upsert_user_meta_for_backup(uid, meta_updates, datetime.now().isoformat())
                        
                        if policy_updates and user_exists:
                            new_policy = {**current_policy, **policy_updates}
                            media_api.post(f"/Users/{uid}/Policy", json=new_policy)
                        
                        success_count += 1
                        results.append({"user_id": uid, "name": uname, "success": True})
                        
                    except Exception as e:
                        fail_count += 1
                        results.append({"user_id": uid, "name": uname, "success": False, "error": str(e)})
                        logger.warning(f"[用户备份] 恢复用户 {uname} 失败: {e}")
                
                # 恢复积分记录（仅覆盖模式）
                if point_logs and mode == "overwrite" and "points" in fields:
                    try:
                        replace_point_logs_for_backup(point_logs)
                        self.log(f"✅ 恢复积分记录 {len(point_logs)} 条")
                    except Exception as e:
                        logger.warning(f"[用户备份] 恢复积分记录失败: {e}")
                
                # 清理临时文件
                if filename.startswith("temp_restore_"):
                    try:
                        os.remove(filepath)
                    except:
                        pass
                
                self.log(f"✅ 恢复完成: 成功 {success_count}, 失败 {fail_count}")
                
                return {
                    "status": "success",
                    "data": {
                        "success_count": success_count,
                        "fail_count": fail_count,
                        "results": results[:50]
                    },
                    "message": f"恢复完成: 成功 {success_count} 个用户，失败 {fail_count} 个"
                }
            except Exception as e:
                logger.error(f"[用户备份] 恢复失败: {e}")
                return {"status": "error", "message": str(e)}

        @self.router.get("/webdav_list")
        async def list_webdav_backups(request: Request):
            """获取 WebDAV 上的备份列表"""
            if not request.session.get("user"):
                return {"status": "error", "message": "未登录"}
            if not is_admin_user(request):
                return {"status": "error", "message": "需要管理员权限"}
            
            if not self._is_pro():
                return {"status": "error", "message": "此功能需要 Pro 授权", "need_pro": True}
            
            webdav = self._get_webdav_config()
            if not webdav["enabled"] or not webdav["url"]:
                return {"status": "error", "message": "WebDAV 未启用"}
            
            try:
                base_url = webdav["url"].rstrip("/")
                path = webdav["path"].lstrip("/")
                if not path.endswith("/"):
                    path += "/"
                dir_url = base_url + "/" + path
                auth = (webdav["user"], webdav["password"]) if webdav["user"] else None
                
                # 使用 PROPFIND 获取目录列表
                resp = webdav_client.request(
                    'PROPFIND',
                    dir_url,
                    auth=auth,
                    headers={'Content-Type': 'application/xml', 'Depth': '1'},
                    data='<?xml version="1.0"?><d:propfind xmlns:d="DAV:"><d:prop><d:displayname/><d:getlastmodified/><d:getcontentlength/></d:prop></d:propfind>',
                    timeout=30,
                )
                
                if resp.status_code not in [200, 207]:
                    return {"status": "error", "message": f"WebDAV 请求失败: HTTP {resp.status_code}"}
                
                # 解析 XML 响应
                import xml.etree.ElementTree as ET
                root = ET.fromstring(resp.content)
                
                backups = []
                for response_elem in root.findall('.//{DAV:}response'):
                    href = response_elem.find('{DAV:}href')
                    if href is None:
                        continue
                    
                    href_text = href.text
                    # 只获取 xlsx 文件
                    if not href_text.endswith('.xlsx') or 'user_backup_' not in href_text:
                        continue
                    
                    # 获取文件名
                    filename = href_text.split('/')[-1]
                    
                    # 获取文件大小和修改时间
                    propstat = response_elem.find('{DAV:}propstat')
                    if propstat:
                        prop = propstat.find('{DAV:}prop')
                        if prop:
                            size_elem = prop.find('{DAV:}getcontentlength')
                            size = int(size_elem.text) if size_elem is not None else 0
                            
                            modified_elem = prop.find('{DAV:}getlastmodified')
                            modified = modified_elem.text if modified_elem is not None else ''
                    
                    backups.append({
                        "filename": filename,
                        "size": size,
                        "modified": modified,
                        "source": "webdav"
                    })
                
                # 按文件名倒序排序（最新的在前）
                backups.sort(key=lambda x: x['filename'], reverse=True)
                
                return {"status": "success", "data": backups}
            except Exception as e:
                logger.error(f"[用户备份] WebDAV 列表失败: {e}")
                return {"status": "error", "message": str(e)}

        @self.router.post("/webdav_download")
        async def download_from_webdav(request: Request):
            """从 WebDAV 下载备份文件"""
            if not request.session.get("user"):
                return {"status": "error", "message": "未登录"}
            if not is_admin_user(request):
                return {"status": "error", "message": "需要管理员权限"}
            
            if not self._is_pro():
                return {"status": "error", "message": "此功能需要 Pro 授权", "need_pro": True}
            
            try:
                data = await request.json()
                filename = data.get("filename")
                
                if not filename:
                    return {"status": "error", "message": "缺少文件名"}
                
                webdav = self._get_webdav_config()
                if not webdav["enabled"] or not webdav["url"]:
                    return {"status": "error", "message": "WebDAV 未启用"}
                
                # 安全检查
                if ".." in filename or "/" in filename or "\\" in filename:
                    return {"status": "error", "message": "无效文件名"}
                
                base_url = webdav["url"].rstrip("/")
                path = webdav["path"].lstrip("/")
                if not path.endswith("/"):
                    path += "/"
                
                file_url = base_url + "/" + path + filename
                auth = (webdav["user"], webdav["password"]) if webdav["user"] else None
                
                # 下载文件
                resp = webdav_client.get(file_url, auth=auth, timeout=60)
                
                if resp.status_code != 200:
                    return {"status": "error", "message": f"下载失败: HTTP {resp.status_code}"}
                
                # 保存到本地
                filepath = os.path.join(BACKUP_DIR, filename)
                with open(filepath, "wb") as f:
                    f.write(resp.content)
                
                file_size = len(resp.content)
                
                return {
                    "status": "success",
                    "data": {
                        "filename": filename,
                        "size": file_size,
                        "local_path": filepath
                    },
                    "message": f"下载成功: {filename} ({file_size} bytes)"
                }
            except Exception as e:
                logger.error(f"[用户备份] WebDAV 下载失败: {e}")
                return {"status": "error", "message": str(e)}

        @self.router.post("/webdav_test")
        async def test_webdav(request: Request):
            """测试 WebDAV 连接"""
            if not request.session.get("user"):
                return {"status": "error", "message": "未登录"}
            if not is_admin_user(request):
                return {"status": "error", "message": "需要管理员权限"}
            
            try:
                data = await request.json()
                url = data.get("url", "")
                user = data.get("user", "")
                password = data.get("password", "")
                
                if not url:
                    return {"status": "error", "message": "请输入 WebDAV 地址"}
                
                auth = (user, password) if user else None
                test_url = url.rstrip("/") + "/"
                
                # 使用 OPTIONS 方法测试 WebDAV 连接（WebDAV 标准方法）
                resp = webdav_client.options(test_url, auth=auth, timeout=10)
                
                # 检查是否支持 WebDAV 方法
                allow_header = resp.headers.get('Allow', '')
                webdav_methods = ['PROPFIND', 'PUT', 'MKCOL', 'DELETE', 'COPY', 'MOVE']
                supports_webdav = any(method in allow_header.upper() for method in webdav_methods)
                
                if resp.status_code in [200, 204, 401]:
                    if resp.status_code == 401:
                        return {"status": "error", "message": "认证失败，请检查用户名密码"}
                    if supports_webdav:
                        return {"status": "success", "message": "WebDAV 连接成功"}
                    else:
                        # 尝试 PROPFIND 方法作为备用测试
                        propfind_resp = webdav_client.request(
                            'PROPFIND',
                            test_url,
                            auth=auth,
                            headers={'Content-Type': 'application/xml', 'Depth': '0'},
                            data='<?xml version="1.0"?><d:propfind xmlns:d="DAV:"><d:prop/></d:propfind>',
                            timeout=10,
                        )
                        if propfind_resp.status_code in [200, 207, 401]:
                            if propfind_resp.status_code == 401:
                                return {"status": "error", "message": "认证失败，请检查用户名密码"}
                            return {"status": "success", "message": "WebDAV 连接成功"}
                        return {"status": "error", "message": "该服务器不支持 WebDAV"}
                else:
                    return {"status": "error", "message": f"连接失败: HTTP {resp.status_code}"}
            except webdav_client.Timeout:
                return {"status": "error", "message": "连接超时"}
            except Exception as e:
                return {"status": "error", "message": str(e)}

    def _cleanup_old_backups(self):
        """清理旧备份"""
        config = self._get_config()
        keep_count = int(config.get("keep_count", 5))
        
        if keep_count <= 0:
            return
        
        try:
            files = sorted([f for f in os.listdir(BACKUP_DIR) 
                          if f.endswith(".xlsx") and f.startswith("user_backup_")], reverse=True)
            
            # 删除超过保留数量的旧文件
            for old_file in files[keep_count:]:
                filepath = os.path.join(BACKUP_DIR, old_file)
                os.remove(filepath)
                logger.info(f"[用户备份] 清理旧备份: {old_file}")
        except Exception as e:
            logger.warning(f"[用户备份] 清理旧备份失败: {e}")

    def on_enable(self):
        self._ensure_dir()
        self._running = True
        # 启动时检查今天是否已有备份文件，避免重启后重复备份
        self._check_today_backup()
        self._thread = threading.Thread(target=self._schedule_loop, daemon=True)
        self._thread.start()
        logger.info("🔌 [用户备份] 插件已启用，定时备份任务已启动")

    def _check_today_backup(self):
        """检查今天是否已有备份文件"""
        try:
            today_str = datetime.now().strftime("%Y%m%d")
            if os.path.exists(BACKUP_DIR):
                for f in os.listdir(BACKUP_DIR):
                    if f.startswith("user_backup_") and f.endswith(".xlsx"):
                        # 文件名格式: user_backup_20260414_124523.xlsx
                        if f[13:21] == today_str:  # 提取日期部分
                            self._last_backup_date = datetime.now().date().isoformat()
                            logger.info(f"[用户备份] 发现今日已有备份: {f}, 跳过启动备份")
                            return
        except Exception as e:
            logger.warning(f"[用户备份] 检查今日备份失败: {e}")

    def on_disable(self):
        self._running = False
        logger.info("🔌 [用户备份] 插件已禁用，定时备份任务已停止")

    def _schedule_loop(self):
        """定时备份循环"""
        time.sleep(30)  # 启动后等待30秒
        
        while self._running and self._enabled:
            try:
                if not self._is_pro():
                    time.sleep(3600)
                    continue
                
                config = self._get_config()
                interval = config.get("backup_interval", "disabled")
                
                if interval == "disabled":
                    # 禁用状态，每小时检查一次配置变更
                    time.sleep(3600)
                    continue
                
                # 解析备份时间
                backup_time_str = config.get("backup_time", "03:00")
                try:
                    hour, minute = backup_time_str.split(":")
                    hour = int(hour)
                    minute = int(minute)
                except:
                    hour, minute = 3, 0
                
                # 计算下次备份时间
                now = datetime.now()
                today_backup = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
                
                # 根据间隔判断是否需要备份
                should_backup = False
                
                # 检查是否在备份时间窗口内（前后5分钟）
                time_diff = (now - today_backup).total_seconds()
                is_in_backup_window = -300 <= time_diff <= 300  # 配置时间前后5分钟
                
                if interval == "daily":
                    # 每日：只在配置时间前后5分钟内备份
                    if self._last_backup_date != now.date().isoformat():
                        if is_in_backup_window:
                            should_backup = True
                
                elif interval == "weekly":
                    # 每周：周一，且在配置时间前后5分钟内
                    if now.weekday() == 0:  # 周一
                        if self._last_backup_date != now.date().isoformat():
                            if is_in_backup_window:
                                should_backup = True
                
                elif interval == "monthly":
                    # 每月：1号，且在配置时间前后5分钟内
                    if now.day == 1:
                        if self._last_backup_date != now.date().isoformat():
                            if is_in_backup_window:
                                should_backup = True
                
                # 执行备份
                if should_backup:
                    self._do_auto_backup()
                    self._last_backup_date = now.date().isoformat()
                    self.log(f"✅ 定时备份完成 ({interval})")
                
                # 等待下一次检查（每分钟检查一次）
                time.sleep(60)
                
            except Exception as e:
                logger.error(f"[用户备份] 定时任务异常: {e}")
                time.sleep(300)

    def _do_auto_backup(self):
        """执行自动备份"""
        try:
            # 创建 Excel
            file_data = self._create_excel()
            
            # 生成文件名
            filename = f"user_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(BACKUP_DIR, filename)
            
            # 保存本地
            with open(filepath, "wb") as f:
                f.write(file_data)
            
            # 上传 WebDAV
            self._upload_to_webdav(filename, file_data)
            
            # 清理旧备份
            self._cleanup_old_backups()
            
            file_size = len(file_data)
            logger.info(f"[用户备份] 定时备份成功: {filename} ({file_size} bytes)")
            
        except Exception as e:
            self.log(f"❌ 定时备份失败: {e}", level="error")

    def get_config_schema(self):
        # 所有配置项已移到面板，不需要单独的配置弹窗
        return []

    def get_page_url(self):
        return "/plugins/user_backup"


# 创建插件实例
user_backup_plugin = UserBackupPlugin()
